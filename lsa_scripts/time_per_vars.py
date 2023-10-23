import subprocess
import time
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.utils import get_column_letter


parser = argparse.ArgumentParser(
    description="Generate and run multiple C++ tests for the static analyzer of C++ Rust-like lifetime annotations."
)
parser.add_argument("-w", "--warnings", default=True, help="count number of warnings")
parser.add_argument("-n", "--notes", default=False, help="count number of notes")
parser.add_argument("-g", "--generate", default=True, help="generate tests")
args = parser.parse_args()


iterations = 10  # Number of iterations for each value of -s
s_values = [0.00, 0.10, 0.20, 0.30, 0.40, 0.5, 0.60, 0.70, 0.80, 0.90, 1]  # List of -s values

size_of_file = 50000

total_tests = iterations * len(s_values)

dir = "gen_tests/"
generator = "../c_code_generator.py"

# Create a new workbook
workbook = Workbook()
data_sheet = workbook.active

# Write headers to the sheet
data_sheet['A1'] = 'File Name'
data_sheet['B1'] = '% Vars'
data_sheet['C1'] = 'Static Analyzer'
data_sheet['D1'] = 'Compile Time'
data_sheet['E1'] = 'Avg Static Analyzer'
data_sheet['F1'] = 'Avg Compile Time'

if args.warnings:
    data_sheet['H1'] = 'Warnings'
if args.notes:
    data_sheet['I1'] = 'Notes'


data_sheet['L1'] = '% Vars'   # 10
data_sheet['M1'] = 'Avg Static Analyzer'
data_sheet['N1'] = 'Avg Compile Time'

if args.warnings:
    data_sheet['Q1'] = 'Avg Warnings'
if args.notes:
    data_sheet['R1'] = 'Avg Notes'

row = 2  # Start writing from row 2
avg_row = 2  # Start writing from row 2

for s in s_values:
    analyzer_times = []
    compile_times = []
    warns = []
    notes = []
    start_row = row  # Starting row for the current s value

    for i in range(1, iterations + 1):
        print(f"[{row - 1}/{total_tests}]")

        filename = f"generated_{int(s*100)}_t{i}"
        file = filename + ".cpp"
        outfile = filename + ".o"
        warnsfile = "warns.txt"

        if args.generate:
            command = f"python {generator} -s {size_of_file} -n {file} -v -i {s} > /dev/null"
        file_path = dir + file
        outfile_path = dir + outfile
        
        # Execute the command using subprocess module
        
        subprocess.run(command, shell=True)

        # Measure time for clang -c -Wprint-lifetimes
        command1 = f"clang -c -Wprint-lifetimes {file_path} -o {outfile_path} 2> /dev/null"
        start_time1 = time.time()
        subprocess.run(command1, shell=True)
        end_time1 = time.time()
        elapsed_time1 = end_time1 - start_time1
        analyzer_times.append(elapsed_time1)
        # print(f"Time taken for clang -c -Wprint-lifetimes {filename}: {elapsed_time1:.2f} seconds")

        if args.warnings:
            commandwarns = f"clang -c -Wprint-lifetimes {file_path} -o {outfile_path} 2> {warnsfile}"
            subprocess.run(commandwarns, shell=True)
            output = subprocess.getoutput(f"grep 'warnings generated' {warnsfile} | grep -oE '[0-9]+'")
            num_warns = int(subprocess.getoutput(f"grep 'warnings generated' {warnsfile} | grep -oE '[0-9]+'"))
            data_sheet.cell(row, 8, value=num_warns)
            warns.append(num_warns)
            if args.notes:
                num_notes = int(subprocess.getoutput(f"grep -c 'note:' {warnsfile}"))
                data_sheet.cell(row, 9, value=num_notes)
                notes.append(num_notes)
        
        # Measure time for clang -c
        command2 = f"clang -c {file_path} -o {outfile_path} 2> /dev/null"
        start_time2 = time.time()
        subprocess.run(command2, shell=True)
        end_time2 = time.time()
        elapsed_time2 = end_time2 - start_time2
        compile_times.append(elapsed_time2)
        # print(f"Time taken for clang -c {filename}: {elapsed_time2:.2f} seconds")

        print(f"File: {file}\tVars: {s}\tProgress: {row - start_row + 1}/{iterations}\t {elapsed_time1:.2f}s")

        # remove .o
        os.remove(outfile_path)

        # Write data to Excel sheet
        data_sheet.cell(row, 1, value=file)
        data_sheet.cell(row, 2, value=s*100)
        data_sheet.cell(row, 3, value=elapsed_time1)
        data_sheet.cell(row, 4, value=elapsed_time2)

        row += 1



    # remove warns file
    os.remove(warnsfile)

    data_sheet.cell(avg_row, 12, value=s*100)

    # Calculate averages
    avg_analyzer_time = sum(analyzer_times) / iterations
    avg_compile_time = sum(compile_times) / iterations

    # Write averages to Excel sheet
    data_sheet.cell(avg_row, 13, value=avg_analyzer_time)
    data_sheet.cell(avg_row, 14, value=avg_compile_time)
    

    if args.warnings:
        avg_warnings = sum(warns) / iterations
        data_sheet.cell(avg_row, 17, value=avg_warnings)
        if args.notes:
            avg_notes = sum(notes) / iterations
            data_sheet.cell(avg_row, 18, value=avg_notes)

    avg_row += 1

    # Write averages to Excel sheet
    data_sheet.cell(row-iterations, 5, value=avg_analyzer_time)
    data_sheet.cell(row-iterations, 6, value=avg_compile_time)

    # Write averages to merged cells
    avg_start_cell = data_sheet.cell(start_row, 5)
    avg_end_cell = data_sheet.cell(row - 1, 5)
    avg_merge_range = f"{avg_start_cell.coordinate}:{avg_end_cell.coordinate}"
    avg_start_cell.value = avg_analyzer_time
    data_sheet.merge_cells(avg_merge_range)
    avg_start_cell.alignment = Alignment(vertical='center')
    
    avg_start_cell = data_sheet.cell(start_row, 6)
    avg_end_cell = data_sheet.cell(row - 1, 6)
    avg_merge_range = f"{avg_start_cell.coordinate}:{avg_end_cell.coordinate}"
    avg_start_cell.value = avg_compile_time
    data_sheet.merge_cells(avg_merge_range)
    avg_start_cell.alignment = Alignment(vertical='center')

# Set column widths
column_widths = [25, 12, 15, 15, 16, 16, 5, 10, 10, 5, 5, 15, 15, 15, 5, 15, 15]  # Adjust the widths as needed

for col_idx, width in enumerate(column_widths, 1):
    col_letter = get_column_letter(col_idx)
    data_sheet.column_dimensions[col_letter].width = width

# Create a new workdata_sheet for the chart
chart_sheet = workbook.create_sheet(title="Chart")

# === Scatter Chart ===
scatter_chart = ScatterChart()
scatter_chart.title = 'Compile Times'
# scatter_chart.style = 13
scatter_chart.x_axis.title = 'Percentage of variable declarations (%)'
scatter_chart.y_axis.title = 'Time (s)'

xvalues = Reference(data_sheet, min_col=12, min_row=2, max_row=avg_row-1)

# compile time
values = Reference(data_sheet, min_col=14, min_row=1, max_row=avg_row-1)
compiler_series = Series(values, xvalues, title_from_data=True)
compiler_series.marker.symbol = "square"
compiler_series.marker.graphicalProperties.solidFill = "3DACCF"
compiler_series.marker.graphicalProperties.line.solidFill = "3DACCF"
compiler_series.graphicalProperties.line.solidFill = "3DACCF"
scatter_chart.series.append(compiler_series)

# static analyzer time
values = Reference(data_sheet, min_col=13, min_row=1, max_row=avg_row-1)
analyzer_series = Series(values, xvalues, title_from_data=True)
analyzer_series.marker.symbol = "circle"
analyzer_series.marker.graphicalProperties.solidFill = "FF9719"
analyzer_series.marker.graphicalProperties.line.solidFill = "FF9719"
analyzer_series.graphicalProperties.line.solidFill = "FF9719"
scatter_chart.series.append(analyzer_series)


chart_sheet.add_chart(scatter_chart, 'H1')

num_rows = len(s_values)*iterations


# === Warnings to Time Chart ===
warns_chart = ScatterChart()
warns_chart.title = 'Compile Times per Warnings and Notes'
warns_chart.x_axis.title = 'Number of Warnings'
warns_chart.y_axis.title = 'Time (s)'
xvalues = Reference(data_sheet, min_col=8, min_row=2, max_row = num_rows + 1)

# time
values = Reference(data_sheet, min_col=3, min_row=1, max_row= num_rows + 1)
warns_series = Series(values, xvalues, title_from_data=True)
warns_series.marker.symbol = "circle"
warns_series.graphicalProperties.line.noFill = True
warns_chart.series.append(warns_series)
chart_sheet.add_chart(warns_chart, 'H20')

workbook.save(f"{dir}compile_times.xlsx")
print("Execution times successfully written to Excel file.")
        
